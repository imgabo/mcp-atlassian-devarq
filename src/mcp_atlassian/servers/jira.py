"""Jira FastMCP server instance and tool definitions."""

import json
import logging
from typing import Annotated, Any
import os
import base64
import shutil
import tempfile
import openpyxl
import csv
import fitz  # PyMuPDF
import re
import requests

from fastmcp import Context, FastMCP
from pydantic import Field
from requests.exceptions import HTTPError

from mcp_atlassian.exceptions import MCPAtlassianAuthenticationError
from mcp_atlassian.jira.constants import DEFAULT_READ_JIRA_FIELDS
from mcp_atlassian.models.jira.common import JiraUser
from mcp_atlassian.utils import convert_empty_defaults_to_none

logger = logging.getLogger(__name__)

jira_mcp = FastMCP(
    name="Jira MCP Service",
    description="Provides tools for interacting with Atlassian Jira.",
)

VOLUME_PATH = "/mnt/archivos"

@jira_mcp.tool(tags={"jira", "read"})
async def list_volume_files(
    ctx: Context,
) -> str:
    """
    List all files in the mounted volume directory (/mnt/archivos).
    Returns a JSON string with a list of relative file paths.
    """
    files = []
    for root, dirs, filenames in os.walk(VOLUME_PATH):
        for filename in filenames:
            rel_path = os.path.relpath(os.path.join(root, filename), VOLUME_PATH)
            files.append(rel_path)
    return json.dumps({"files": files}, indent=2, ensure_ascii=False)

@jira_mcp.tool(tags={"jira", "read"})
async def read_volume_file(
    ctx: Context,
    filename: Annotated[str, Field(description="Relative path of the file inside /mnt/archivos")],
) -> str:
    """
    Read the content of a file in the mounted volume directory (/mnt/archivos).
    Returns the content as text if possible, or as base64 if binary. Si es PDF, extrae el texto.
    """
    file_path = os.path.abspath(os.path.join(VOLUME_PATH, filename))
    if not file_path.startswith(VOLUME_PATH):
        return json.dumps({"error": "Invalid file path."}, indent=2, ensure_ascii=False)
    if not os.path.exists(file_path):
        return json.dumps({"error": "File not found."}, indent=2, ensure_ascii=False)
    # Si es PDF, extraer texto
    if file_path.lower().endswith(".pdf"):
        try:
            import PyPDF2
            with open(file_path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() or ""
            return json.dumps({"filename": filename, "content": text}, indent=2, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"error": f"Failed to extract PDF text: {str(e)}"}, indent=2, ensure_ascii=False)
    # Si es texto plano
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        return json.dumps({"filename": filename, "content": content}, indent=2, ensure_ascii=False)
    except UnicodeDecodeError:
        with open(file_path, "rb") as f:
            content = base64.b64encode(f.read()).decode("utf-8")
        return json.dumps({"filename": filename, "content_base64": content}, indent=2, ensure_ascii=False)

@jira_mcp.tool(tags={"jira", "read"})
async def get_user_profile(
    ctx: Context,
    user_identifier: Annotated[
        str,
        Field(
            description="Identifier for the user (e.g., email address 'user@example.com', username 'johndoe', account ID 'accountid:...', or key for Server/DC)."
        ),
    ],
) -> str:
    """
    Retrieve profile information for a specific Jira user.

    Args:
        ctx: The FastMCP context.
        user_identifier: User identifier (email, username, key, or account ID).

    Returns:
        JSON string representing the Jira user profile object, or an error object if not found.

    Raises:
        ValueError: If the Jira client is not configured or available.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    try:
        user: JiraUser = jira.get_user_profile_by_identifier(user_identifier)
        result = user.to_simplified_dict()
        response_data = {"success": True, "user": result}
    except Exception as e:
        error_message = ""
        log_level = logging.ERROR

        if isinstance(e, ValueError) and "not found" in str(e).lower():
            log_level = logging.WARNING
            error_message = str(e)
        elif isinstance(e, MCPAtlassianAuthenticationError):
            error_message = f"Authentication/Permission Error: {str(e)}"
        elif isinstance(e, OSError | HTTPError):
            error_message = f"Network or API Error: {str(e)}"
        else:
            error_message = (
                "An unexpected error occurred while fetching the user profile."
            )
            logger.exception(
                f"Unexpected error in get_user_profile for '{user_identifier}':"
            )

        error_result = {
            "success": False,
            "error": str(e),
            "user_identifier": user_identifier,
        }
        logger.log(
            log_level,
            f"get_user_profile failed for '{user_identifier}': {error_message}",
        )
        response_data = error_result

    return json.dumps(response_data, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "read"})
async def get_issue(
    ctx: Context,
    issue_key: Annotated[str, Field(description="Jira issue key (e.g., 'PROJ-123')")],
    fields: Annotated[
        str,
        Field(
            description=(
                "(Optional) Comma-separated list of fields to return (e.g., 'summary,status,customfield_10010'). "
                "You may also provide a single field as a string (e.g., 'duedate'). "
                "Use '*all' for all fields (including custom fields), or omit for essential fields only."
            ),
            default=",".join(DEFAULT_READ_JIRA_FIELDS),
        ),
    ] = ",".join(DEFAULT_READ_JIRA_FIELDS),
    expand: Annotated[
        str,
        Field(
            description=(
                "(Optional) Fields to expand. Examples: 'renderedFields' (for rendered content), "
                "'transitions' (for available status transitions), 'changelog' (for history)"
            ),
            default="",
        ),
    ] = "",
    comment_limit: Annotated[
        int,
        Field(
            description="Maximum number of comments to include (0 or null for no comments)",
            default=10,
            ge=0,
            le=100,
        ),
    ] = 10,
    properties: Annotated[
        str,
        Field(
            description="(Optional) A comma-separated list of issue properties to return",
            default="",
        ),
    ] = "",
    update_history: Annotated[
        bool,
        Field(
            description="Whether to update the issue view history for the requesting user",
            default=True,
        ),
    ] = True,
) -> str:
    """Get details of a specific Jira issue including its Epic links and relationship information.

    Args:
        ctx: The FastMCP context.
        issue_key: Jira issue key.
        fields: Comma-separated list of fields to return (e.g., 'summary,status,customfield_10010'), a single field as a string (e.g., 'duedate'), '*all' for all fields, or omitted for essentials.
        expand: Optional fields to expand.
        comment_limit: Maximum number of comments.
        properties: Issue properties to return.
        update_history: Whether to update issue view history.

    Returns:
        JSON string representing the Jira issue object.

    Raises:
        ValueError: If the Jira client is not configured or available.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    fields_list: str | list[str] | None = fields
    if fields and fields != "*all":
        fields_list = [f.strip() for f in fields.split(",")]

    issue = jira.get_issue(
        issue_key=issue_key,
        fields=fields_list,
        expand=expand,
        comment_limit=comment_limit,
        properties=properties.split(",") if properties else None,
        update_history=update_history,
    )
    result = issue.to_simplified_dict()
    return json.dumps(result, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "read"})
async def search(
    ctx: Context,
    jql: Annotated[
        str,
        Field(
            description=(
                "JQL query string (Jira Query Language). Examples:\n"
                '- Find Epics: "issuetype = Epic AND project = PROJ"\n'
                '- Find issues in Epic: "parent = PROJ-123"\n'
                "- Find by status: \"status = 'In Progress' AND project = PROJ\"\n"
                '- Find by assignee: "assignee = currentUser()"\n'
                '- Find recently updated: "updated >= -7d AND project = PROJ"\n'
                '- Find by label: "labels = frontend AND project = PROJ"\n'
                '- Find by priority: "priority = High AND project = PROJ"'
            )
        ),
    ],
    fields: Annotated[
        str,
        Field(
            description=(
                "(Optional) Comma-separated fields to return in the results. "
                "Use '*all' for all fields, or specify individual fields like 'summary,status,assignee,priority'"
            ),
            default=",".join(DEFAULT_READ_JIRA_FIELDS),
        ),
    ] = ",".join(DEFAULT_READ_JIRA_FIELDS),
    limit: Annotated[
        int,
        Field(description="Maximum number of results (1-50)", default=10, ge=1),
    ] = 10,
    start_at: Annotated[
        int,
        Field(description="Starting index for pagination (0-based)", default=0, ge=0),
    ] = 0,
    projects_filter: Annotated[
        str,
        Field(
            description=(
                "(Optional) Comma-separated list of project keys to filter results by. "
                "Overrides the environment variable JIRA_PROJECTS_FILTER if provided."
            ),
            default="",
        ),
    ] = "",
    expand: Annotated[
        str,
        Field(
            description=(
                "(Optional) fields to expand. Examples: 'renderedFields', 'transitions', 'changelog'"
            ),
            default="",
        ),
    ] = "",
) -> str:
    """Search Jira issues using JQL (Jira Query Language).

    Args:
        ctx: The FastMCP context.
        jql: JQL query string.
        fields: Comma-separated fields to return.
        limit: Maximum number of results.
        start_at: Starting index for pagination.
        projects_filter: Comma-separated list of project keys to filter by.
        expand: Optional fields to expand.

    Returns:
        JSON string representing the search results including pagination info.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    fields_list: str | list[str] | None = fields
    if fields and fields != "*all":
        fields_list = [f.strip() for f in fields.split(",")]

    search_result = jira.search_issues(
        jql=jql,
        fields=fields_list,
        limit=limit,
        start=start_at,
        expand=expand,
        projects_filter=projects_filter,
    )
    result = search_result.to_simplified_dict()
    return json.dumps(result, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "read"})
async def search_fields(
    ctx: Context,
    keyword: Annotated[
        str,
        Field(
            description="Keyword for fuzzy search. If left empty, lists the first 'limit' available fields in their default order.",
            default="",
        ),
    ] = "",
    limit: Annotated[
        int, Field(description="Maximum number of results", default=10, ge=1)
    ] = 10,
    refresh: Annotated[
        bool,
        Field(description="Whether to force refresh the field list", default=False),
    ] = False,
) -> str:
    """Search Jira fields by keyword with fuzzy match.

    Args:
        ctx: The FastMCP context.
        keyword: Keyword for fuzzy search.
        limit: Maximum number of results.
        refresh: Whether to force refresh the field list.

    Returns:
        JSON string representing a list of matching field definitions.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    result = jira.search_fields(keyword, limit=limit, refresh=refresh)
    return json.dumps(result, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "read"})
async def get_project_issues(
    ctx: Context,
    project_key: Annotated[str, Field(description="The project key")],
    limit: Annotated[
        int,
        Field(description="Maximum number of results (1-50)", default=10, ge=1, le=50),
    ] = 10,
    start_at: Annotated[
        int,
        Field(description="Starting index for pagination (0-based)", default=0, ge=0),
    ] = 0,
) -> str:
    """Get all issues for a specific Jira project.

    Args:
        ctx: The FastMCP context.
        project_key: The project key.
        limit: Maximum number of results.
        start_at: Starting index for pagination.

    Returns:
        JSON string representing the search results including pagination info.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    search_result = jira.get_project_issues(
        project_key=project_key, start=start_at, limit=limit
    )
    result = search_result.to_simplified_dict()
    return json.dumps(result, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "read"})
async def get_transitions(
    ctx: Context,
    issue_key: Annotated[str, Field(description="Jira issue key (e.g., 'PROJ-123')")],
) -> str:
    """Get available status transitions for a Jira issue.

    Args:
        ctx: The FastMCP context.
        issue_key: Jira issue key.

    Returns:
        JSON string representing a list of available transitions.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    # Underlying method returns list[dict] in the desired format
    transitions = jira.get_available_transitions(issue_key)
    return json.dumps(transitions, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "read"})
async def get_worklog(
    ctx: Context,
    issue_key: Annotated[str, Field(description="Jira issue key (e.g., 'PROJ-123')")],
) -> str:
    """Get worklog entries for a Jira issue.

    Args:
        ctx: The FastMCP context.
        issue_key: Jira issue key.

    Returns:
        JSON string representing the worklog entries.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    worklogs = jira.get_worklogs(issue_key)
    result = {"worklogs": worklogs}
    return json.dumps(result, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "read"})
async def download_attachments(
    ctx: Context,
    issue_key: Annotated[str, Field(description="Jira issue key (e.g., 'PROJ-123')")],
    target_dir: Annotated[
        str, Field(description="Directory where attachments should be saved")
    ],
) -> str:
    """Download attachments from a Jira issue.

    Args:
        ctx: The FastMCP context.
        issue_key: Jira issue key.
        target_dir: Directory to save attachments.

    Returns:
        JSON string indicating the result of the download operation.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    result = jira.download_issue_attachments(issue_key=issue_key, target_dir=target_dir)
    return json.dumps(result, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "read"})
async def get_agile_boards(
    ctx: Context,
    board_name: Annotated[
        str, Field(description="(Optional) The name of board, support fuzzy search")
    ] = "",
    project_key: Annotated[
        str, Field(description="(Optional) Jira project key (e.g., 'PROJ-123')")
    ] = "",
    board_type: Annotated[
        str,
        Field(
            description="(Optional) The type of jira board (e.g., 'scrum', 'kanban')"
        ),
    ] = "",
    start_at: Annotated[
        int,
        Field(description="Starting index for pagination (0-based)", default=0, ge=0),
    ] = 0,
    limit: Annotated[
        int,
        Field(description="Maximum number of results (1-50)", default=10, ge=1, le=50),
    ] = 10,
) -> str:
    """Get jira agile boards by name, project key, or type.

    Args:
        ctx: The FastMCP context.
        board_name: Name of the board (fuzzy search).
        project_key: Project key.
        board_type: Board type ('scrum' or 'kanban').
        start_at: Starting index.
        limit: Maximum results.

    Returns:
        JSON string representing a list of board objects.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    boards = jira.get_all_agile_boards_model(
        board_name=board_name,
        project_key=project_key,
        board_type=board_type,
        start=start_at,
        limit=limit,
    )
    result = [board.to_simplified_dict() for board in boards]
    return json.dumps(result, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "read"})
async def get_board_issues(
    ctx: Context,
    board_id: Annotated[str, Field(description="The id of the board (e.g., '1001')")],
    jql: Annotated[
        str,
        Field(
            description=(
                "JQL query string (Jira Query Language). Examples:\n"
                '- Find Epics: "issuetype = Epic AND project = PROJ"\n'
                '- Find issues in Epic: "parent = PROJ-123"\n'
                "- Find by status: \"status = 'In Progress' AND project = PROJ\"\n"
                '- Find by assignee: "assignee = currentUser()"\n'
                '- Find recently updated: "updated >= -7d AND project = PROJ"\n'
                '- Find by label: "labels = frontend AND project = PROJ"\n'
                '- Find by priority: "priority = High AND project = PROJ"'
            )
        ),
    ],
    fields: Annotated[
        str,
        Field(
            description=(
                "Comma-separated fields to return in the results. "
                "Use '*all' for all fields, or specify individual "
                "fields like 'summary,status,assignee,priority'"
            ),
            default=",".join(DEFAULT_READ_JIRA_FIELDS),
        ),
    ] = ",".join(DEFAULT_READ_JIRA_FIELDS),
    start_at: Annotated[
        int,
        Field(description="Starting index for pagination (0-based)", default=0, ge=0),
    ] = 0,
    limit: Annotated[
        int,
        Field(description="Maximum number of results (1-50)", default=10, ge=1, le=50),
    ] = 10,
    expand: Annotated[
        str,
        Field(
            description="Optional fields to expand in the response (e.g., 'changelog').",
            default="version",
        ),
    ] = "version",
) -> str:
    """Get all issues linked to a specific board filtered by JQL.

    Args:
        ctx: The FastMCP context.
        board_id: The ID of the board.
        jql: JQL query string to filter issues.
        fields: Comma-separated fields to return.
        start_at: Starting index for pagination.
        limit: Maximum number of results.
        expand: Optional fields to expand.

    Returns:
        JSON string representing the search results including pagination info.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    fields_list: str | list[str] | None = fields
    if fields and fields != "*all":
        fields_list = [f.strip() for f in fields.split(",")]

    search_result = jira.get_board_issues(
        board_id=board_id,
        jql=jql,
        fields=fields_list,
        start=start_at,
        limit=limit,
        expand=expand,
    )
    result = search_result.to_simplified_dict()
    return json.dumps(result, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "read"})
async def get_sprints_from_board(
    ctx: Context,
    board_id: Annotated[str, Field(description="The id of board (e.g., '1000')")],
    state: Annotated[
        str,
        Field(description="Sprint state (e.g., 'active', 'future', 'closed')"),
    ] = "",
    start_at: Annotated[
        int,
        Field(description="Starting index for pagination (0-based)", default=0, ge=0),
    ] = 0,
    limit: Annotated[
        int,
        Field(description="Maximum number of results (1-50)", default=10, ge=1, le=50),
    ] = 10,
) -> str:
    """Get jira sprints from board by state.

    Args:
        ctx: The FastMCP context.
        board_id: The ID of the board.
        state: Sprint state ('active', 'future', 'closed'). If None, returns all sprints.
        start_at: Starting index.
        limit: Maximum results.

    Returns:
        JSON string representing a list of sprint objects.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    sprints = jira.get_all_sprints_from_board_model(
        board_id=board_id, state=state, start=start_at, limit=limit
    )
    result = [sprint.to_simplified_dict() for sprint in sprints]
    return json.dumps(result, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "read"})
async def get_sprint_issues(
    ctx: Context,
    sprint_id: Annotated[str, Field(description="The id of sprint (e.g., '10001')")],
    fields: Annotated[
        str,
        Field(
            description=(
                "Comma-separated fields to return in the results. "
                "Use '*all' for all fields, or specify individual "
                "fields like 'summary,status,assignee,priority'"
            ),
            default=",".join(DEFAULT_READ_JIRA_FIELDS),
        ),
    ] = ",".join(DEFAULT_READ_JIRA_FIELDS),
    start_at: Annotated[
        int,
        Field(description="Starting index for pagination (0-based)", default=0, ge=0),
    ] = 0,
    limit: Annotated[
        int,
        Field(description="Maximum number of results (1-50)", default=10, ge=1, le=50),
    ] = 10,
) -> str:
    """Get jira issues from sprint.

    Args:
        ctx: The FastMCP context.
        sprint_id: The ID of the sprint.
        fields: Comma-separated fields to return.
        start_at: Starting index.
        limit: Maximum results.

    Returns:
        JSON string representing the search results including pagination info.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    fields_list: str | list[str] | None = fields
    if fields and fields != "*all":
        fields_list = [f.strip() for f in fields.split(",")]

    search_result = jira.get_sprint_issues(
        sprint_id=sprint_id, fields=fields_list, start=start_at, limit=limit
    )
    result = search_result.to_simplified_dict()
    return json.dumps(result, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "read"})
async def get_link_types(ctx: Context) -> str:
    """Get all available issue link types.

    Args:
        ctx: The FastMCP context.

    Returns:
        JSON string representing a list of issue link type objects.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    link_types = jira.get_issue_link_types()
    formatted_link_types = [link_type.to_simplified_dict() for link_type in link_types]
    return json.dumps(formatted_link_types, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "write"})
async def create_issue(
    ctx: Context,
    project_key: Annotated[
        str,
        Field(
            description=(
                "The JIRA project key (e.g. 'PROJ', 'DEV', 'SUPPORT'). "
                "This is the prefix of issue keys in your project. "
                "Never assume what it might be, always ask the user."
            )
        ),
    ],
    summary: Annotated[str, Field(description="Summary/title of the issue")],
    issue_type: Annotated[
        str,
        Field(
            description=(
                "Issue type (e.g. 'Task', 'Bug', 'Story', 'Epic', 'Subtask'). "
                "The available types depend on your project configuration. "
                "For subtasks, use 'Subtask' (not 'Sub-task') and include parent in additional_fields."
            )
        ),
    ],
    assignee: Annotated[
        str,
        Field(
            description="(Optional) Assignee's user identifier (string): Email, display name, or account ID (e.g., 'user@example.com', 'John Doe', 'accountid:...')",
            default="",
        ),
    ] = "",
    description: Annotated[
        str, Field(description="Issue description", default="")
    ] = "",
    components: Annotated[
        str,
        Field(
            description="(Optional) Comma-separated list of component names to assign (e.g., 'Frontend,API')",
            default="",
        ),
    ] = "",
    additional_fields: Annotated[
        dict[str, Any],
        Field(
            description=(
                "(Optional) Dictionary of additional fields to set. Examples:\n"
                "- Set priority: {'priority': {'name': 'High'}}\n"
                "- Add labels: {'labels': ['frontend', 'urgent']}\n"
                "- Link to parent (for any issue type): {'parent': 'PROJ-123'}\n"
                "- Set Fix Version/s: {'fixVersions': [{'id': '10020'}]}\n"
                "- Custom fields: {'customfield_10010': 'value'}"
            ),
            default_factory=dict,
        ),
    ] = {},  # noqa: B006
) -> str:
    """Create a new Jira issue with optional Epic link or parent for subtasks.

    Args:
        ctx: The FastMCP context.
        project_key: The JIRA project key.
        summary: Summary/title of the issue.
        issue_type: Issue type (e.g., 'Task', 'Bug', 'Story', 'Epic', 'Subtask').
        assignee: Assignee's user identifier (string): Email, display name, or account ID (e.g., 'user@example.com', 'John Doe', 'accountid:...').
        description: Issue description.
        components: Comma-separated list of component names.
        additional_fields: Dictionary of additional fields.

    Returns:
        JSON string representing the created issue object.

    Raises:
        ValueError: If in read-only mode or Jira client is unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call create_issue in read-only mode.")
        raise ValueError("Cannot create issue in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    # Parse components from comma-separated string to list
    components_list = None
    if components and isinstance(components, str):
        components_list = [
            comp.strip() for comp in components.split(",") if comp.strip()
        ]

    # Use additional_fields directly as dict
    extra_fields = additional_fields or {}
    if not isinstance(extra_fields, dict):
        raise ValueError("additional_fields must be a dictionary.")

    # Convert description to Wiki Markup (Jira format)
    description_wiki = jira.markdown_to_jira(description) if description else ""

    issue = jira.create_issue(
        project_key=project_key,
        summary=summary,
        issue_type=issue_type,
        description=description_wiki,
        assignee=assignee,
        components=components_list,
        **extra_fields,
    )
    result = issue.to_simplified_dict()
    return json.dumps(
        {"message": "Issue created successfully", "issue": result},
        indent=2,
        ensure_ascii=False,
    )


@jira_mcp.tool(tags={"jira", "write"})
async def batch_create_issues(
    ctx: Context,
    issues: Annotated[
        str,
        Field(
            description=(
                "JSON array of issue objects. Each object should contain:\n"
                "- project_key (required): The project key (e.g., 'PROJ')\n"
                "- summary (required): Issue summary/title\n"
                "- issue_type (required): Type of issue (e.g., 'Task', 'Bug')\n"
                "- description (optional): Issue description\n"
                "- assignee (optional): Assignee username or email\n"
                "- components (optional): Array of component names\n"
                "Example: [\n"
                '  {"project_key": "PROJ", "summary": "Issue 1", "issue_type": "Task"},\n'
                '  {"project_key": "PROJ", "summary": "Issue 2", "issue_type": "Bug", "components": ["Frontend"]}\n'
                "]"
            )
        ),
    ],
    validate_only: Annotated[
        bool,
        Field(
            description="If true, only validates the issues without creating them",
            default=False,
        ),
    ] = False,
) -> str:
    """Create multiple Jira issues in a batch.

    Args:
        ctx: The FastMCP context.
        issues: JSON array string of issue objects.
        validate_only: If true, only validates without creating.

    Returns:
        JSON string indicating success and listing created issues (or validation result).

    Raises:
        ValueError: If in read-only mode, Jira client unavailable, or invalid JSON.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call batch_create_issues in read-only mode.")
        raise ValueError("Cannot create issues in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    # Parse issues from JSON string
    try:
        issues_list = json.loads(issues)
        if not isinstance(issues_list, list):
            raise ValueError("Input 'issues' must be a JSON array string.")
    except json.JSONDecodeError:
        raise ValueError("Invalid JSON in issues")
    except Exception as e:
        raise ValueError(f"Invalid input for issues: {e}") from e

    # Create issues in batch
    created_issues = jira.batch_create_issues(issues_list, validate_only=validate_only)

    message = (
        "Issues validated successfully"
        if validate_only
        else "Issues created successfully"
    )
    result = {
        "message": message,
        "issues": [issue.to_simplified_dict() for issue in created_issues],
    }
    return json.dumps(result, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "read"})
async def batch_get_changelogs(
    ctx: Context,
    issue_ids_or_keys: Annotated[
        list[str],
        Field(
            description="List of Jira issue IDs or keys, e.g. ['PROJ-123', 'PROJ-124']"
        ),
    ],
    fields: Annotated[
        list[str],
        Field(
            description="(Optional) Filter the changelogs by fields, e.g. ['status', 'assignee']. Default to [] for all fields.",
            default_factory=list,
        ),
    ] = [],  # noqa: B006
    limit: Annotated[
        int,
        Field(
            description=(
                "Maximum number of changelogs to return in result for each issue. "
                "Default to -1 for all changelogs. "
                "Notice that it only limits the results in the response, "
                "the function will still fetch all the data."
            ),
            default=-1,
        ),
    ] = -1,
) -> str:
    """Get changelogs for multiple Jira issues (Cloud only).

    Args:
        ctx: The FastMCP context.
        issue_ids_or_keys: List of issue IDs or keys.
        fields: List of fields to filter changelogs by. None for all fields.
        limit: Maximum changelogs per issue (-1 for all).

    Returns:
        JSON string representing a list of issues with their changelogs.

    Raises:
        NotImplementedError: If run on Jira Server/Data Center.
        ValueError: If Jira client is unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    # Ensure this runs only on Cloud, as per original function docstring
    if not jira.config.is_cloud:
        raise NotImplementedError(
            "Batch get issue changelogs is only available on Jira Cloud."
        )

    # Call the underlying method
    issues_with_changelogs = jira.batch_get_changelogs(
        issue_ids_or_keys=issue_ids_or_keys, fields=fields
    )

    # Format the response
    results = []
    limit_val = None if limit == -1 else limit
    for issue in issues_with_changelogs:
        results.append(
            {
                "issue_id": issue.id,
                "changelogs": [
                    changelog.to_simplified_dict()
                    for changelog in issue.changelogs[:limit_val]
                ],
            }
        )
    return json.dumps(results, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "write"})
async def update_issue(
    ctx: Context,
    issue_key: Annotated[str, Field(description="Jira issue key (e.g., 'PROJ-123')")],
    fields: Annotated[
        dict[str, Any],
        Field(
            description=(
                "Dictionary of fields to update. For 'assignee', provide a string identifier (email, name, or accountId). "
                "Example: `{'assignee': 'user@example.com', 'summary': 'New Summary'}`"
            )
        ),
    ],
    additional_fields: Annotated[
        dict[str, Any],
        Field(
            description="(Optional) Dictionary of additional fields to update. Use this for custom fields or more complex updates.",
            default_factory=dict,
        ),
    ] = {},  # noqa: B006
    attachments: Annotated[
        str,
        Field(
            description=(
                "(Optional) JSON string array or comma-separated list of file paths to attach to the issue. "
                "Example: '/path/to/file1.txt,/path/to/file2.txt' or ['/path/to/file1.txt','/path/to/file2.txt']"
            ),
            default="",
        ),
    ] = "",
) -> str:
    """Update an existing Jira issue including changing status, adding Epic links, updating fields, etc.

    Args:
        ctx: The FastMCP context.
        issue_key: Jira issue key.
        fields: Dictionary of fields to update.
        additional_fields: Optional dictionary of additional fields.
        attachments: Optional JSON array string or comma-separated list of file paths.

    Returns:
        JSON string representing the updated issue object and attachment results.

    Raises:
        ValueError: If in read-only mode, Jira client unavailable, or invalid input.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call update_issue in read-only mode.")
        raise ValueError("Cannot update issue in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    # Use fields directly as dict
    if not isinstance(fields, dict):
        raise ValueError("fields must be a dictionary.")
    update_fields = fields

    # Use additional_fields directly as dict
    extra_fields = additional_fields or {}
    if not isinstance(extra_fields, dict):
        raise ValueError("additional_fields must be a dictionary.")

    # Parse attachments
    attachment_paths = []
    if attachments:
        if isinstance(attachments, str):
            try:
                parsed = json.loads(attachments)
                if isinstance(parsed, list):
                    attachment_paths = [str(p) for p in parsed]
                else:
                    raise ValueError("attachments JSON string must be an array.")
            except json.JSONDecodeError:
                # Assume comma-separated if not valid JSON array
                attachment_paths = [
                    p.strip() for p in attachments.split(",") if p.strip()
                ]
        else:
            raise ValueError(
                "attachments must be a JSON array string or comma-separated string."
            )

    # Combine fields and additional_fields
    all_updates = {**update_fields, **extra_fields}
    if attachment_paths:
        all_updates["attachments"] = attachment_paths

    try:
        issue = jira.update_issue(issue_key=issue_key, **all_updates)
        result = issue.to_simplified_dict()
        if (
            hasattr(issue, "custom_fields")
            and "attachment_results" in issue.custom_fields
        ):
            result["attachment_results"] = issue.custom_fields["attachment_results"]
        return json.dumps(
            {"message": "Issue updated successfully", "issue": result},
            indent=2,
            ensure_ascii=False,
        )
    except Exception as e:
        logger.error(f"Error updating issue {issue_key}: {str(e)}", exc_info=True)
        raise ValueError(f"Failed to update issue {issue_key}: {str(e)}")


@jira_mcp.tool(tags={"jira", "write"})
async def delete_issue(
    ctx: Context,
    issue_key: Annotated[str, Field(description="Jira issue key (e.g. PROJ-123)")],
) -> str:
    """Delete an existing Jira issue.

    Args:
        ctx: The FastMCP context.
        issue_key: Jira issue key.

    Returns:
        JSON string indicating success.

    Raises:
        ValueError: If in read-only mode or Jira client unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call delete_issue in read-only mode.")
        raise ValueError("Cannot delete issue in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    deleted = jira.delete_issue(issue_key)
    result = {"message": f"Issue {issue_key} has been deleted successfully."}
    # The underlying method raises on failure, so if we reach here, it's success.
    return json.dumps(result, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "write"})
async def add_comment(
    ctx: Context,
    issue_key: Annotated[str, Field(description="Jira issue key (e.g., 'PROJ-123')")],
    comment: Annotated[str, Field(description="Comment text in Markdown format")],
) -> str:
    """Add a comment to a Jira issue.

    Args:
        ctx: The FastMCP context.
        issue_key: Jira issue key.
        comment: Comment text in Markdown.

    Returns:
        JSON string representing the added comment object.

    Raises:
        ValueError: If in read-only mode or Jira client unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call add_comment in read-only mode.")
        raise ValueError("Cannot add comment in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    # add_comment returns dict
    result = jira.add_comment(issue_key, comment)
    return json.dumps(result, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "write"})
async def add_worklog(
    ctx: Context,
    issue_key: Annotated[str, Field(description="Jira issue key (e.g., 'PROJ-123')")],
    time_spent: Annotated[
        str,
        Field(
            description=(
                "Time spent in Jira format. Examples: "
                "'1h 30m' (1 hour and 30 minutes), '1d' (1 day), '30m' (30 minutes), '4h' (4 hours)"
            )
        ),
    ],
    comment: Annotated[
        str,
        Field(description="(Optional) Comment for the worklog in Markdown format"),
    ] = "",
    started: Annotated[
        str,
        Field(
            description=(
                "(Optional) Start time in ISO format. If not provided, the current time will be used. "
                "Example: '2023-08-01T12:00:00.000+0000'"
            )
        ),
    ] = "",
    # Add original_estimate and remaining_estimate as per original tool
    original_estimate: Annotated[
        str, Field(description="(Optional) New value for the original estimate")
    ] = "",
    remaining_estimate: Annotated[
        str, Field(description="(Optional) New value for the remaining estimate")
    ] = "",
) -> str:
    """Add a worklog entry to a Jira issue.

    Args:
        ctx: The FastMCP context.
        issue_key: Jira issue key.
        time_spent: Time spent in Jira format.
        comment: Optional comment in Markdown.
        started: Optional start time in ISO format.
        original_estimate: Optional new original estimate.
        remaining_estimate: Optional new remaining estimate.


    Returns:
        JSON string representing the added worklog object.

    Raises:
        ValueError: If in read-only mode or Jira client unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call add_worklog in read-only mode.")
        raise ValueError("Cannot add worklog in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    # add_worklog returns dict
    worklog_result = jira.add_worklog(
        issue_key=issue_key,
        time_spent=time_spent,
        comment=comment,
        started=started,
        original_estimate=original_estimate,
        remaining_estimate=remaining_estimate,
    )
    result = {"message": "Worklog added successfully", "worklog": worklog_result}
    return json.dumps(result, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "write"})
async def link_to_epic(
    ctx: Context,
    issue_key: Annotated[
        str, Field(description="The key of the issue to link (e.g., 'PROJ-123')")
    ],
    epic_key: Annotated[
        str, Field(description="The key of the epic to link to (e.g., 'PROJ-456')")
    ],
) -> str:
    """Link an existing issue to an epic.

    Args:
        ctx: The FastMCP context.
        issue_key: The key of the issue to link.
        epic_key: The key of the epic to link to.

    Returns:
        JSON string representing the updated issue object.

    Raises:
        ValueError: If in read-only mode or Jira client unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call link_to_epic in read-only mode.")
        raise ValueError("Cannot link issue to epic in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    issue = jira.link_issue_to_epic(issue_key, epic_key)
    result = {
        "message": f"Issue {issue_key} has been linked to epic {epic_key}.",
        "issue": issue.to_simplified_dict(),
    }
    return json.dumps(result, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "write"})
async def create_issue_link(
    ctx: Context,
    link_type: Annotated[
        str,
        Field(
            description="The type of link to create (e.g., 'Duplicate', 'Blocks', 'Relates to')"
        ),
    ],
    inward_issue_key: Annotated[
        str, Field(description="The key of the inward issue (e.g., 'PROJ-123')")
    ],
    outward_issue_key: Annotated[
        str, Field(description="The key of the outward issue (e.g., 'PROJ-456')")
    ],
    comment: Annotated[
        str, Field(description="(Optional) Comment to add to the link")
    ] = "",
    comment_visibility: Annotated[
        dict[str, str],
        Field(
            description="(Optional) Visibility settings for the comment (e.g., {'type': 'group', 'value': 'jira-users'})",
            default_factory=dict,
        ),
    ] = {},  # noqa: B006
) -> str:
    """Create a link between two Jira issues.

    Args:
        ctx: The FastMCP context.
        link_type: The type of link (e.g., 'Blocks').
        inward_issue_key: The key of the source issue.
        outward_issue_key: The key of the target issue.
        comment: Optional comment text.
        comment_visibility: Optional dictionary for comment visibility.

    Returns:
        JSON string indicating success or failure.

    Raises:
        ValueError: If required fields are missing, invalid input, in read-only mode, or Jira client unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call create_issue_link in read-only mode.")
        raise ValueError("Cannot create issue link in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    if not all([link_type, inward_issue_key, outward_issue_key]):
        raise ValueError(
            "link_type, inward_issue_key, and outward_issue_key are required."
        )

    link_data = {
        "type": {"name": link_type},
        "inwardIssue": {"key": inward_issue_key},
        "outwardIssue": {"key": outward_issue_key},
    }

    if comment:
        comment_obj = {"body": comment}
        if comment_visibility and isinstance(comment_visibility, dict):
            if "type" in comment_visibility and "value" in comment_visibility:
                comment_obj["visibility"] = comment_visibility
            else:
                logger.warning("Invalid comment_visibility dictionary structure.")
        link_data["comment"] = comment_obj

    result = jira.create_issue_link(link_data)
    return json.dumps(result, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "write"})
async def remove_issue_link(
    ctx: Context,
    link_id: Annotated[str, Field(description="The ID of the link to remove")],
) -> str:
    """Remove a link between two Jira issues.

    Args:
        ctx: The FastMCP context.
        link_id: The ID of the link to remove.

    Returns:
        JSON string indicating success.

    Raises:
        ValueError: If link_id is missing, in read-only mode, or Jira client unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call remove_issue_link in read-only mode.")
        raise ValueError("Cannot remove issue link in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    if not link_id:
        raise ValueError("link_id is required")

    result = jira.remove_issue_link(link_id)  # Returns dict on success
    return json.dumps(result, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "write"})
async def transition_issue(
    ctx: Context,
    issue_key: Annotated[str, Field(description="Jira issue key (e.g., 'PROJ-123')")],
    transition_id: Annotated[
        str,
        Field(
            description=(
                "ID of the transition to perform. Use the jira_get_transitions tool first "
                "to get the available transition IDs for the issue. Example values: '11', '21', '31'"
            )
        ),
    ],
    fields: Annotated[
        dict[str, Any],
        Field(
            description=(
                "(Optional) Dictionary of fields to update during the transition. "
                "Some transitions require specific fields to be set (e.g., resolution). "
                "Example: {'resolution': {'name': 'Fixed'}}"
            ),
            default_factory=dict,
        ),
    ] = {},  # noqa: B006
    comment: Annotated[
        str,
        Field(
            description=(
                "(Optional) Comment to add during the transition. "
                "This will be visible in the issue history."
            ),
        ),
    ] = "",
) -> str:
    """Transition a Jira issue to a new status.

    Args:
        ctx: The FastMCP context.
        issue_key: Jira issue key.
        transition_id: ID of the transition.
        fields: Optional dictionary of fields to update during transition.
        comment: Optional comment for the transition.

    Returns:
        JSON string representing the updated issue object.

    Raises:
        ValueError: If required fields missing, invalid input, in read-only mode, or Jira client unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call transition_issue in read-only mode.")
        raise ValueError("Cannot transition issue in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    if not issue_key or not transition_id:
        raise ValueError("issue_key and transition_id are required.")

    # Use fields directly as dict
    update_fields = fields or {}
    if not isinstance(update_fields, dict):
        raise ValueError("fields must be a dictionary.")

    issue = jira.transition_issue(
        issue_key=issue_key,
        transition_id=transition_id,
        fields=update_fields,
        comment=comment,
    )

    result = {
        "message": f"Issue {issue_key} transitioned successfully",
        "issue": issue.to_simplified_dict() if issue else None,
    }
    return json.dumps(result, indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "write"})
async def create_sprint(
    ctx: Context,
    board_id: Annotated[str, Field(description="The id of board (e.g., '1000')")],
    sprint_name: Annotated[
        str, Field(description="Name of the sprint (e.g., 'Sprint 1')")
    ],
    start_date: Annotated[
        str, Field(description="Start time for sprint (ISO 8601 format)")
    ],
    end_date: Annotated[
        str, Field(description="End time for sprint (ISO 8601 format)")
    ],
    goal: Annotated[str, Field(description="(Optional) Goal of the sprint")] = "",
) -> str:
    """Create Jira sprint for a board.

    Args:
        ctx: The FastMCP context.
        board_id: Board ID.
        sprint_name: Sprint name.
        start_date: Start date (ISO format).
        end_date: End date (ISO format).
        goal: Optional sprint goal.

    Returns:
        JSON string representing the created sprint object.

    Raises:
        ValueError: If in read-only mode or Jira client unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call create_sprint in read-only mode.")
        raise ValueError("Cannot create sprint in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    sprint = jira.create_sprint(
        board_id=board_id,
        sprint_name=sprint_name,
        start_date=start_date,
        end_date=end_date,
        goal=goal,
    )
    return json.dumps(sprint.to_simplified_dict(), indent=2, ensure_ascii=False)


@convert_empty_defaults_to_none
@jira_mcp.tool(tags={"jira", "write"})
async def update_sprint(
    ctx: Context,
    sprint_id: Annotated[str, Field(description="The id of sprint (e.g., '10001')")],
    sprint_name: Annotated[
        str, Field(description="(Optional) New name for the sprint")
    ] = "",
    state: Annotated[
        str,
        Field(description="(Optional) New state for the sprint (future|active|closed)"),
    ] = "",
    start_date: Annotated[
        str, Field(description="(Optional) New start date for the sprint")
    ] = "",
    end_date: Annotated[
        str, Field(description="(Optional) New end date for the sprint")
    ] = "",
    goal: Annotated[str, Field(description="(Optional) New goal for the sprint")] = "",
) -> str:
    """Update jira sprint.

    Args:
        ctx: The FastMCP context.
        sprint_id: The ID of the sprint.
        sprint_name: Optional new name.
        state: Optional new state (future|active|closed).
        start_date: Optional new start date.
        end_date: Optional new end date.
        goal: Optional new goal.

    Returns:
        JSON string representing the updated sprint object or an error message.

    Raises:
        ValueError: If in read-only mode or Jira client unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call update_sprint in read-only mode.")
        raise ValueError("Cannot update sprint in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    sprint = jira.update_sprint(
        sprint_id=sprint_id,
        sprint_name=sprint_name,
        state=state,
        start_date=start_date,
        end_date=end_date,
        goal=goal,
    )

    if sprint is None:
        error_payload = {
            "error": f"Failed to update sprint {sprint_id}. Check logs for details."
        }
        return json.dumps(error_payload, indent=2, ensure_ascii=False)
    else:
        return json.dumps(sprint.to_simplified_dict(), indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "write"})
async def upload_attachment(
    ctx: Context,
    issue_key: Annotated[str, Field(description="Jira issue key (e.g., 'PROJ-123')")],
    file_path: Annotated[
        str,
        Field(
            description=(
                "Path to the file to upload OR base64 encoded image data starting with 'data:image/'"
            )
        ),
    ],
) -> str:
    """Upload a single attachment to a Jira issue.

    Args:
        ctx: The FastMCP context.
        issue_key: Jira issue key.
        file_path: Path to the file to upload OR base64 encoded image data starting with 'data:image/'.

    Returns:
        JSON string indicating the result of the upload operation.

    Raises:
        ValueError: If in read-only mode or Jira client unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call upload_attachment in read-only mode.")
        raise ValueError("Cannot upload attachment in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    # Check if the input is a base64 image data
    if file_path.startswith('data:image/'):
        try:
            # Extract the base64 data and file extension
            header, encoded = file_path.split(",", 1)
            file_extension = header.split(";")[0].split("/")[1]
            
            # Create a temporary file with the image data
            import base64
            import tempfile
            import os
            
            # Decode base64 data
            image_data = base64.b64decode(encoded)
            
            # Create a temporary file with the correct extension
            with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_extension}') as temp_file:
                temp_file.write(image_data)
                temp_file_path = temp_file.name
            
            try:
                # Upload the temporary file
                result = jira.upload_attachment(issue_key=issue_key, file_path=temp_file_path)
                return json.dumps(result, indent=2, ensure_ascii=False)
            finally:
                # Clean up the temporary file
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
        except Exception as e:
            raise ValueError(f"Failed to process base64 image data: {str(e)}")
    else:
        # Handle regular file path
        result = jira.upload_attachment(issue_key=issue_key, file_path=file_path)
        return json.dumps(result, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "write"})
async def upload_attachment_with_comment(
    ctx: Context,
    issue_key: Annotated[str, Field(description="Jira issue key (e.g., 'PROJ-123')")],
    file_path: Annotated[
        str,
        Field(
            description=(
                "Path to the file to upload OR base64 encoded image data starting with 'data:image/'"
            )
        ),
    ],
    comment: Annotated[
        str,
        Field(
            description="(Optional) Additional comment to include with the attachment.",
            default="",
        ),
    ] = "",
) -> str:
    """
    Upload a single attachment to a Jira issue and add a comment referencing the file (with Wiki Markup for embedding).

    Args:
        ctx: The FastMCP context.
        issue_key: Jira issue key.
        file_path: Path to the file to upload OR base64 encoded image data starting with 'data:image/'.
        comment: (Optional) Additional comment to include.

    Returns:
        JSON string indicating the result of the upload and comment operation.

    Raises:
        ValueError: If in read-only mode or Jira client unavailable.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call upload_attachment_with_comment in read-only mode.")
        raise ValueError("Cannot upload attachment or add comment in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    # 1. Subir el attachment (reutiliza la lógica de upload_attachment)
    if file_path.startswith('data:image/'):
        try:
            header, encoded = file_path.split(",", 1)
            file_extension = header.split(";")[0].split("/")[1]
            import base64
            import tempfile
            import os
            image_data = base64.b64decode(encoded)
            with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_extension}') as temp_file:
                temp_file.write(image_data)
                temp_file_path = temp_file.name
            try:
                upload_result = jira.upload_attachment(issue_key=issue_key, file_path=temp_file_path)
            finally:
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
        except Exception as e:
            raise ValueError(f"Failed to process base64 image data: {str(e)}")
    else:
        upload_result = jira.upload_attachment(issue_key=issue_key, file_path=file_path)

    # Si falla el upload, retorna el error
    if not upload_result or not upload_result.get("success"):
        return json.dumps({
            "success": False,
            "error": upload_result.get("error", "Unknown error uploading attachment"),
            "stage": "upload_attachment"
        }, indent=2, ensure_ascii=False)

    # 2. Agregar comentario mencionando el archivo con Wiki Markup
    filename = upload_result.get("filename") or file_path.split("/")[-1]
    comment_text = f"Imagen adjunta: {filename}\n!{filename}!"
    if comment:
        comment_text += f"\n{comment}"
    comment_result = jira.add_comment(issue_key, comment_text)

    # Si falla el comentario, retorna el error pero indica que el archivo sí fue subido
    if not comment_result or not comment_result.get("id"):
        return json.dumps({
            "success": False,
            "error": "Attachment uploaded but failed to add comment.",
            "attachment": upload_result,
            "stage": "add_comment"
        }, indent=2, ensure_ascii=False)

    # Éxito total
    return json.dumps({
        "success": True,
        "attachment": upload_result,
        "comment": comment_result
    }, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "write"})
async def bulk_delete_comments(
    ctx: Context,
    issue_key: Annotated[str, Field(description="Jira issue key (e.g., 'PROJ-123')")],
    comment_ids: Annotated[list[str], Field(description="List of comment IDs to delete")],
) -> str:
    """
    Delete multiple comments from a Jira issue.

    Args:
        ctx: The FastMCP context.
        issue_key: Jira issue key.
        comment_ids: List of comment IDs to delete.

    Returns:
        JSON string summarizing which comments were deleted and which failed.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        logger.warning("Attempted to call bulk_delete_comments in read-only mode.")
        raise ValueError("Cannot delete comments in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    deleted = []
    failed = []
    for comment_id in comment_ids:
        success = jira.delete_comment(issue_key, comment_id)
        if success:
            deleted.append(comment_id)
        else:
            failed.append(comment_id)

    result = {
        "success": len(failed) == 0,
        "deleted": deleted,
        "failed": failed,
        "issue_key": issue_key,
    }
    return json.dumps(result, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "write"})
async def create_issue_from_volume_file(
    ctx: Context,
    project_key: Annotated[str, Field(description="Jira project key (e.g., 'PROJ')")],
    issue_type: Annotated[str, Field(description="Issue type (e.g., 'Task', 'Story')")],
    filename: Annotated[str, Field(description="Relative path of the file inside /mnt/archivos")],
    assignee: Annotated[str, Field(description="(Optional) Assignee's user identifier", default="")] = "",
    components: Annotated[str, Field(description="(Optional) Comma-separated list of component names", default="")] = "",
) -> str:
    """
    Crea un issue en Jira usando el contenido de un archivo del volumen y adjunta el archivo y sus imágenes embebidas.
    Soporta .docx, .pdf y texto plano. Extrae imágenes y las referencia en la descripción.
    """
    import os
    import base64
    import shutil
    import tempfile
    VOLUME_PATH = "/mnt/archivos"
    file_path = os.path.abspath(os.path.join(VOLUME_PATH, filename))
    if not file_path.startswith(VOLUME_PATH):
        return json.dumps({"success": False, "error": "Invalid file path."}, indent=2, ensure_ascii=False)
    if not os.path.exists(file_path):
        return json.dumps({"success": False, "error": "File not found."}, indent=2, ensure_ascii=False)

    # Variables para resultado
    summary = ""
    description_parts = []
    image_files = []
    temp_dir = tempfile.mkdtemp(dir=VOLUME_PATH)
    error = None

    try:
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".docx":
            from docx import Document
            from docx.document import Document as _Document
            from docx.table import Table
            from docx.text.paragraph import Paragraph
            def iter_block_items(parent):
                if isinstance(parent, _Document):
                    parent_elm = parent.element.body
                else:
                    parent_elm = parent._tc
                for child in parent_elm.iterchildren():
                    if child.tag.endswith('}p'):
                        yield Paragraph(child, parent)
                    elif child.tag.endswith('}tbl'):
                        yield Table(child, parent)
            doc = Document(file_path)
            img_count = 0
            rels = {r.rId: r for r in doc.part.rels.values() if "image" in r.target_ref}
            img_map = {}
            # Mapear imágenes a su nombre temporal
            for rId, rel in rels.items():
                img_count += 1
                img_data = rel.target_part.blob
                img_name = f"{os.path.splitext(os.path.basename(filename))[0]}_img{img_count}.png"
                img_path = os.path.join(temp_dir, img_name)
                with open(img_path, "wb") as img_file:
                    img_file.write(img_data)
                img_map[rId] = img_name
                image_files.append(img_path)
            # Recorrer bloques en orden
            for block in iter_block_items(doc):
                if isinstance(block, Paragraph):
                    text = block.text.strip()
                    if text:
                        description_parts.append(text)
                    # Buscar imágenes en los runs
                    for run in block.runs:
                        if run.element.xpath('.//a:blip'):
                            for blip in run.element.xpath('.//a:blip'):
                                rEmbed = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                if rEmbed and rEmbed in img_map:
                                    description_parts.append(f"!{img_map[rEmbed]}!")
                elif isinstance(block, Table):
                    description_parts.append("[Tabla detectada]")
            summary = description_parts[0] if description_parts else os.path.basename(filename)
            description = "\n\n".join(description_parts[1:]) if len(description_parts) > 1 else ""
        elif ext == ".pdf":
            doc = fitz.open(file_path)
            img_count = 0
            text_parts = []
            image_refs = []
            for page_num, page in enumerate(doc, 1):
                text = page.get_text().strip()
                if text:
                    text_parts.append(text)
                for img_index, img in enumerate(page.get_images(full=True)):
                    xref = img[0]
                    base_img = doc.extract_image(xref)
                    img_bytes = base_img["image"]
                    img_ext = base_img["ext"]
                    img_name = f"{os.path.splitext(os.path.basename(filename))[0]}_p{page_num}_img{img_index+1}.{img_ext}"
                    img_path = os.path.join(temp_dir, img_name)
                    with open(img_path, "wb") as img_file:
                        img_file.write(img_bytes)
                    image_files.append(img_path)
                    image_refs.append(f"!{img_name}!")
            summary = text_parts[0].splitlines()[0] if text_parts and text_parts[0] else os.path.basename(filename)
            description = "\n\n".join(text_parts)
            # Agregar referencias a imágenes al final
            if image_refs:
                description += "\n\n" + "\n\n".join(image_refs)
        elif ext in [".txt", ".md", ".csv"]:
            with open(file_path, "r", encoding="utf-8") as f:
                lines = f.readlines()
            summary = lines[0].strip() if lines else os.path.basename(filename)
            description = "".join(lines[1:]).strip() if len(lines) > 1 else ""
        elif ext == ".xlsx":
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                sheet_texts = []
                for sheet in wb.worksheets:
                    rows = list(sheet.iter_rows(values_only=True))
                    if not rows:
                        continue
                    # Encabezados
                    headers = [str(h) if h is not None else "" for h in rows[0]]
                    # Filas
                    table_lines = ["| " + " | ".join(headers) + " |"]
                    table_lines.append("|" + "|".join(["---"] * len(headers)) + "|")
                    for row in rows[1:]:
                        row_cells = [str(cell) if cell is not None else "" for cell in row]
                        table_lines.append("| " + " | ".join(row_cells) + " |")
                    sheet_text = f"### Hoja: {sheet.title}\n" + "\n".join(table_lines)
                    sheet_texts.append(sheet_text)
                summary = os.path.basename(filename)
                description = "\n\n".join(sheet_texts)
            except Exception as e:
                raise Exception(f"Error al leer archivo .xlsx: {e}")
        elif ext in [".csv", ".xlsx"]:
            # Heurística de mapeo de campos
            FIELD_KEYWORDS = {
                "summary": ["nombre hu", "título", "summary", "resumen", "titulo"],
                "assignee": ["asignado", "responsable", "assignee"],
                "priority": ["prioridad", "priority"],
                "labels": ["etiqueta", "labels"],
                "duedate": ["fecha de vencimiento", "due date", "vencimiento"],
                "storypoints": ["story point", "puntos", "puntos de historia"],
                "status": ["estado", "status"],
            }
            def guess_jira_field(header):
                header_lower = header.strip().lower()
                for jira_field, keywords in FIELD_KEYWORDS.items():
                    if any(kw in header_lower for kw in keywords):
                        return jira_field
                return None
            # Leer datos
            rows = []
            headers = []
            if ext == ".csv":
                with open(file_path, "r", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for i, row in enumerate(reader):
                        if i == 0:
                            headers = [str(h).strip() for h in row]
                        else:
                            rows.append(row)
            else:  # .xlsx
                wb = openpyxl.load_workbook(file_path, data_only=True)
                sheet = wb.active
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if i == 0:
                        headers = [str(h).strip() if h is not None else "" for h in row]
                    else:
                        rows.append([str(cell) if cell is not None else "" for cell in row])
            # Mapeo de columnas
            col_map = {}
            for idx, h in enumerate(headers):
                field = guess_jira_field(h)
                if field:
                    col_map[field] = idx
            if "summary" not in col_map:
                return json.dumps({"success": False, "error": "No se encontró columna de resumen (summary) en el archivo."}, indent=2, ensure_ascii=False)
            # Preparar Jira client
            lifespan_ctx = ctx.request_context.lifespan_context
            if not lifespan_ctx or not lifespan_ctx.jira:
                raise Exception("Jira client is not configured or available.")
            jira = lifespan_ctx.jira
            results = []
            for row in rows:
                try:
                    summary = row[col_map["summary"]].strip() if row[col_map["summary"]] else ""
                    if not summary:
                        continue
                    assignee = row[col_map["assignee"]].strip() if "assignee" in col_map and row[col_map["assignee"]] else ""
                    priority = row[col_map["priority"]].strip() if "priority" in col_map and row[col_map["priority"]] else ""
                    labels = row[col_map["labels"]].strip() if "labels" in col_map and row[col_map["labels"]] else ""
                    duedate = row[col_map["duedate"]].strip() if "duedate" in col_map and row[col_map["duedate"]] else ""
                    storypoints = row[col_map["storypoints"]].strip() if "storypoints" in col_map and row[col_map["storypoints"]] else ""
                    # status no se puede setear directo al crear, pero se puede guardar para referencia
                    additional_fields = {}
                    if priority:
                        additional_fields["priority"] = {"name": priority}
                    if labels:
                        additional_fields["labels"] = [l.strip() for l in labels.split(",") if l.strip()]
                    if duedate:
                        additional_fields["duedate"] = duedate
                    if storypoints and storypoints.isdigit():
                        additional_fields["customfield_10016"] = int(storypoints)  # Ajusta el ID según tu Jira
                    issue = jira.create_issue(
                        project_key=project_key,
                        summary=summary,
                        issue_type=issue_type,
                        assignee=assignee,
                        description="",
                        **additional_fields
                    )
                    results.append({"success": True, "issue": issue.to_simplified_dict() if hasattr(issue, "to_simplified_dict") else issue, "row": summary})
                except Exception as e:
                    results.append({"success": False, "error": str(e), "row": row})
            return json.dumps({"results": results}, indent=2, ensure_ascii=False)
        elif ext in [".txt", ".md"]:
            with open(file_path, "r", encoding="utf-8") as f:
                lines = f.readlines()
            summary = lines[0].strip() if lines else os.path.basename(filename)
            description = "".join(lines[1:]).strip() if len(lines) > 1 else ""
        else:
            raise Exception("Formato de archivo no soportado para extracción automática. Usa .docx, .pdf, .txt, .md, .csv, .xlsx.")
        # Construir descripción final (texto + imágenes)
        if description_parts:
            description = "\n\n".join(description_parts)
        # Convertir a Wiki Markup
        lifespan_ctx = ctx.request_context.lifespan_context
        if not lifespan_ctx or not lifespan_ctx.jira:
            raise Exception("Jira client is not configured or available.")
        jira = lifespan_ctx.jira
        description_wiki = jira.markdown_to_jira(description) if description else ""
        # Crear issue
        components_list = [c.strip() for c in components.split(",") if c.strip()] if components else None
        issue = jira.create_issue(
            project_key=project_key,
            summary=summary,
            issue_type=issue_type,
            description=description_wiki,
            assignee=assignee,
            components=components_list,
        )
        issue_key = getattr(issue, "key", None) or (issue.get("key") if isinstance(issue, dict) else None)
        if not issue_key:
            raise Exception("Failed to create issue.")
        # Subir attachments (imágenes extraídas)
        attachment_results = []
        for img_path in image_files:
            result = jira.upload_attachment(issue_key=issue_key, file_path=img_path)
            attachment_results.append(result)
        # Limpiar archivos temporales
        shutil.rmtree(temp_dir)
        return json.dumps({
            "success": True,
            "issue": issue.to_simplified_dict() if hasattr(issue, "to_simplified_dict") else issue,
            "attachments": attachment_results
        }, indent=2, ensure_ascii=False)
    except Exception as e:
        error = str(e)
        # Limpiar archivos temporales si hubo error
        try:
            shutil.rmtree(temp_dir)
        except Exception:
            pass
        return json.dumps({"success": False, "error": error}, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "write"})
async def create_issue_with_secondary_activities(
    ctx: Context,
    project_key: Annotated[str, Field(description="Jira project key (e.g., 'PROJ')")],
    summary: Annotated[str, Field(description="Summary/title of the main issue")],
    issue_type: Annotated[str, Field(description="Issue type for the main issue (e.g., 'Story', 'Task')")],
    description: Annotated[str, Field(description="Description of the main issue", default="")] = "",
    assignee: Annotated[str, Field(description="(Optional) Assignee for the main issue", default="")] = "",
    components: Annotated[str, Field(description="(Optional) Comma-separated list of component names", default="")] = "",
    additional_fields: Annotated[
        dict[str, Any],
        Field(description="(Optional) Dictionary of additional fields for the main issue", default_factory=dict),
    ] = {},
    secondary_activities: Annotated[
        list[dict],
        Field(description="(Optional) List of dicts with all fields for each secondary activity/subtask", default_factory=list),
    ] = [],
) -> str:
    """
    Crea un issue principal (HU, Task, etc.) y, opcionalmente, una lista de subtareas (actividades secundarias),
    cada una aceptando todos los campos posibles.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        raise ValueError("Cannot create issue in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    # Preparar componentes
    components_list = [c.strip() for c in components.split(",") if c.strip()] if components else None
    # Convertir descripción a Wiki Markup
    description_wiki = jira.markdown_to_jira(description) if description else ""
    # Crear issue principal
    main_issue = jira.create_issue(
        project_key=project_key,
        summary=summary,
        issue_type=issue_type,
        description=description_wiki,
        assignee=assignee,
        components=components_list,
        **(additional_fields or {})
    )
    main_issue_key = getattr(main_issue, "key", None) or (main_issue.get("key") if isinstance(main_issue, dict) else None)
    if not main_issue_key:
        raise Exception("Failed to create main issue.")
    # Crear subtareas
    created_subtasks = []
    for sub in secondary_activities:
        sub_fields = dict(sub)  # Copia para no mutar el original
        sub_fields["issue_type"] = sub_fields.get("issue_type", "Subtask")
        sub_fields["parent"] = main_issue_key
        # Eliminar campos que no son de Jira.create_issue
        project_key_sub = sub_fields.pop("project_key", project_key)
        summary_sub = sub_fields.pop("summary", None)
        if not summary_sub:
            continue  # No crear si no hay summary
        description_sub = sub_fields.pop("description", "")
        description_sub_wiki = jira.markdown_to_jira(description_sub) if description_sub else ""
        assignee_sub = sub_fields.pop("assignee", "")
        components_sub = sub_fields.pop("components", "")
        components_list_sub = [c.strip() for c in components_sub.split(",") if c.strip()] if components_sub else None
        # Crear subtarea
        subtask = jira.create_issue(
            project_key=project_key_sub,
            summary=summary_sub,
            issue_type=sub_fields.pop("issue_type", "Subtask"),
            description=description_sub_wiki,
            assignee=assignee_sub,
            components=components_list_sub,
            parent=main_issue_key,
            **sub_fields
        )
        created_subtasks.append(subtask.to_simplified_dict() if hasattr(subtask, "to_simplified_dict") else subtask)
    return json.dumps({
        "success": True,
        "main_issue": main_issue.to_simplified_dict() if hasattr(main_issue, "to_simplified_dict") else main_issue,
        "secondary_activities": created_subtasks
    }, indent=2, ensure_ascii=False)


@jira_mcp.tool(tags={"jira", "write"})
async def add_secondary_activity_to_issue(
    ctx: Context,
    parent_issue_key: Annotated[str, Field(description="Issue key of the parent (e.g., HU) to which subtasks will be added")],
    secondary_activities: Annotated[
        list[dict],
        Field(description="List of dicts with all fields for each secondary activity/subtask to add. Cada subtarea debe incluir obligatoriamente los campos 'project_key' y 'summary'.", default_factory=list),
    ] = [],
) -> str:
    """
    Agrega una o más subtareas (actividades secundarias) a un issue existente, aceptando todos los campos posibles.
    Cada subtarea debe incluir obligatoriamente los campos 'project_key' y 'summary'.
    Si falta alguno de estos campos en alguna subtarea, se arrojará un error.
    """
    lifespan_ctx = ctx.request_context.lifespan_context
    if lifespan_ctx.read_only:
        raise ValueError("Cannot create subtasks in read-only mode.")
    if not lifespan_ctx or not lifespan_ctx.jira:
        raise ValueError("Jira client is not configured or available.")
    jira = lifespan_ctx.jira

    if not secondary_activities or not isinstance(secondary_activities, list):
        raise ValueError("Debe proporcionar al menos una subtarea en 'secondary_activities'.")

    created_subtasks = []
    for idx, sub in enumerate(secondary_activities):
        sub_fields = dict(sub)
        project_key_sub = sub_fields.get("project_key")
        summary_sub = sub_fields.get("summary")
        if not project_key_sub or not summary_sub:
            raise ValueError(f"Cada subtarea debe incluir los campos obligatorios 'project_key' y 'summary'. Faltan en la subtarea de índice {idx}.")
        sub_fields["issue_type"] = sub_fields.get("issue_type", "Subtask")
        # NO agregar sub_fields["parent"] = parent_issue_key
        # Extraer y preparar campos
        project_key_sub = sub_fields.pop("project_key")
        summary_sub = sub_fields.pop("summary")
        description_sub = sub_fields.pop("description", "")
        description_sub_wiki = jira.markdown_to_jira(description_sub) if description_sub else ""
        assignee_sub = sub_fields.pop("assignee", "")
        components_sub = sub_fields.pop("components", "")
        components_list_sub = [c.strip() for c in components_sub.split(",") if components_sub and c.strip()] if components_sub else None
        subtask = jira.create_issue(
            project_key=project_key_sub,
            summary=summary_sub,
            issue_type=sub_fields.pop("issue_type", "Subtask"),
            description=description_sub_wiki,
            assignee=assignee_sub,
            components=components_list_sub,
            parent=parent_issue_key,
            **sub_fields
        )
        created_subtasks.append(subtask.to_simplified_dict() if hasattr(subtask, "to_simplified_dict") else subtask)
    return json.dumps({
        "success": True,
        "parent_issue_key": parent_issue_key,
        "secondary_activities": created_subtasks
    }, indent=2, ensure_ascii=False)
#   